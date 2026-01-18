from __future__ import annotations

import io
import logging
import mimetypes
from datetime import datetime, timezone
from typing import Any, cast
from uuid import UUID

from app.core.storage import FileCategory, FileScope, generate_storage_key, get_storage_service
from app.infra.database import AsyncSessionLocal
from app.models.file import FileCreate
from app.repos.file import FileRepository
from app.repos.knowledge_set import KnowledgeSetRepository


logger = logging.getLogger(__name__)


def sanitize_topic(topic: str | None) -> str:
    s = (topic or "literature").strip().lower()
    out: list[str] = []
    for ch in s:
        if ch.isalnum():
            out.append(ch)
        elif ch in {" ", "_", "-", "."}:
            out.append("-")
    sanitized = "".join(out).strip("-")
    return sanitized or "literature"


def derive_xlsx_filename(topic: str | None) -> str:
    ts = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    return f"{sanitize_topic(topic)}-{ts}.xlsx"


def _extract_first_author_affiliation(raw: dict[str, Any] | None) -> str:
    if not isinstance(raw, dict):
        return ""
    # OpenAlex style
    try:
        auths = raw.get("authorships")
        if isinstance(auths, list) and auths:
            insts = auths[0].get("institutions")
            if isinstance(insts, list) and insts:
                name = insts[0].get("display_name") or insts[0].get("name")
                if isinstance(name, str):
                    return name
    except Exception:
        pass
    # Crossref style
    try:
        authors = raw.get("author")
        if isinstance(authors, list) and authors:
            affs = authors[0].get("affiliation")
            if isinstance(affs, list) and affs:
                name = affs[0].get("name")
                if isinstance(name, str):
                    return name
    except Exception:
        pass
    # Semantic Scholar style
    try:
        authors = raw.get("authors")
        if isinstance(authors, list) and authors:
            affs = authors[0].get("affiliations")
            if isinstance(affs, list) and affs:
                name = affs[0].get("name")
                if isinstance(name, str):
                    return name
    except Exception:
        pass
    return ""


def _extract_reference_fields(raw: dict[str, Any] | None) -> dict[str, str]:
    if not isinstance(raw, dict):
        return {"journal": "", "publisher": "", "volume": "", "issue": "", "pages": "", "published_date": ""}
    journal = ""
    publisher = ""
    volume = ""
    issue = ""
    pages = ""
    published_date = ""
    # OpenAlex
    try:
        hv = raw.get("host_venue") or {}
        if isinstance(hv, dict):
            jn = hv.get("display_name") or hv.get("name")
            if isinstance(jn, str):
                journal = jn
            pub = hv.get("publisher")
            if isinstance(pub, str):
                publisher = pub
        biblio = raw.get("biblio") or {}
        if isinstance(biblio, dict):
            vol = biblio.get("volume")
            iss = biblio.get("issue")
            fp = biblio.get("first_page")
            lp = biblio.get("last_page")
            if isinstance(vol, (str, int)):
                volume = str(vol)
            if isinstance(iss, (str, int)):
                issue = str(iss)
            fp_s = str(fp) if isinstance(fp, (str, int)) else ""
            lp_s = str(lp) if isinstance(lp, (str, int)) else ""
            if fp_s or lp_s:
                pages = f"{fp_s}-{lp_s}".strip("-")
        pd = raw.get("publication_date") or raw.get("published_date")
        if isinstance(pd, str):
            published_date = pd
    except Exception:
        pass
    # Crossref
    try:
        cr = raw.get("container-title")
        if isinstance(cr, list) and cr:
            journal = journal or (cr[0] if isinstance(cr[0], str) else "")
        pub = raw.get("publisher")
        if isinstance(pub, str):
            publisher = publisher or pub
        vol = raw.get("volume")
        iss = raw.get("issue")
        pg = raw.get("page")
        if isinstance(vol, (str, int)):
            volume = volume or str(vol)
        if isinstance(iss, (str, int)):
            issue = issue or str(iss)
        if isinstance(pg, str):
            pages = pages or pg
        # date-parts
        issued = raw.get("issued") or {}
        parts_any = issued.get("date-parts")
        if isinstance(parts_any, list) and parts_any and isinstance(parts_any[0], list):
            first = cast(list[Any], parts_any[0])
            y = first[0] if len(first) > 0 else None
            m = first[1] if len(first) > 1 else None
            d = first[2] if len(first) > 2 else None
            published_date = published_date or "-".join([str(x) for x in [y, m, d] if x is not None])
    except Exception:
        pass
    return {
        "journal": journal or "",
        "publisher": publisher or "",
        "volume": volume or "",
        "issue": issue or "",
        "pages": pages or "",
        "published_date": published_date or "",
    }


def build_xlsx_bytes(payload: dict[str, Any], query: str | None, title_hint: str | None) -> bytes:
    """Build an XLSX file (bytes) for literature export.

    We intentionally use XLSX to avoid CSV encoding issues for non-English characters.
    """

    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for XLSX export. Please install 'openpyxl'.")

    wb = openpyxl.Workbook()
    ws_summary = wb.active
    if ws_summary is None:
        ws_summary = wb.create_sheet(title="Summary")
    else:
        ws_summary.title = "Summary"

    meta_any = payload.get("meta")
    meta = cast(dict[str, Any], meta_any) if isinstance(meta_any, dict) else {}
    stats_any = payload.get("stats")
    stats = cast(dict[str, Any], stats_any) if isinstance(stats_any, dict) else {}
    created_at = payload.get("created_at")
    results = cast(list[dict[str, Any]], payload.get("results") or [])

    # Summary
    ws_summary.append(["Query", query or ""])
    ws_summary.append(["TitleHint", title_hint or ""])
    ws_summary.append(
        [
            "Providers",
            ",".join(meta.get("providers", meta.get("requested_providers", [])) or []),
        ]
    )
    ws_summary.append(["Fetched", len(results)])
    ws_summary.append(["CreatedAt", str(created_at or datetime.now(timezone.utc).isoformat())])

    if stats:
        ws_summary.append([])
        ws_summary.append(["Provider", "requests", "fetched", "took_ms"])
        for pname, pstats in stats.items():
            if isinstance(pstats, dict):
                ws_summary.append(
                    [
                        str(pname),
                        pstats.get("requests"),
                        pstats.get("fetched"),
                        pstats.get("took_ms"),
                    ]
                )

    # Works
    ws_works = wb.create_sheet(title="Works")
    ws_works.append(
        [
            "source",
            "source_id",
            "doi",
            "title",
            "authors",
            "year",
            "venue",
            "url",
            "pdf_url",
            "first_author",
            "first_author_affiliation",
            "journal",
            "publisher",
            "volume",
            "issue",
            "pages",
            "published_date",
        ]
    )

    for w in results:
        authors = cast(list[dict[str, Any]], w.get("authors") or [])
        names: list[str] = []
        fa_name = ""
        for a in authors:
            nm = a.get("name")
            if isinstance(nm, str) and nm:
                names.append(nm)
        if names:
            fa_name = names[0]
        authors_str = "; ".join(names)
        raw = cast(dict[str, Any] | None, w.get("raw"))
        first_affil = _extract_first_author_affiliation(raw)
        ref = _extract_reference_fields(raw)
        ws_works.append(
            [
                w.get("source") or "",
                w.get("source_id") or "",
                w.get("doi") or "",
                w.get("title") or "",
                authors_str,
                w.get("year") or "",
                w.get("venue") or "",
                w.get("url") or "",
                w.get("pdf_url") or "",
                fa_name,
                first_affil,
                ref.get("journal") or "",
                ref.get("publisher") or "",
                ref.get("volume") or "",
                ref.get("issue") or "",
                ref.get("pages") or "",
                ref.get("published_date") or "",
            ]
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def persist_xlsx(user_id: str, knowledge_set_id: UUID, filename: str, content: bytes) -> dict[str, Any]:
    try:
        filename = filename.strip("/").split("/")[-1]

        async with AsyncSessionLocal() as db:
            file_repo = FileRepository(db)
            ks_repo = KnowledgeSetRepository(db)
            storage = get_storage_service()

            # Validate access and get file IDs (match knowledge.py behavior)
            try:
                await ks_repo.validate_access(user_id, knowledge_set_id)
                file_ids = await ks_repo.get_files_in_knowledge_set(knowledge_set_id)
            except ValueError as e:
                return {"error": f"Access denied: {e}", "success": False}

            # Determine content type (match knowledge.py fallback style)
            content_type, _ = mimetypes.guess_type(filename)
            if not content_type:
                if filename.endswith(".docx"):
                    content_type = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                elif filename.endswith(".xlsx"):
                    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                elif filename.endswith(".pptx"):
                    content_type = "application/vnd.openxmlformats-officedocument.presentationml.presentation"
                elif filename.endswith(".pdf"):
                    content_type = "application/pdf"
                else:
                    content_type = "text/plain"

            new_key = generate_storage_key(user_id, filename, FileScope.PRIVATE)
            data = io.BytesIO(content)
            file_size_bytes = len(content)
            await storage.upload_file(data, new_key, content_type=content_type)

            # Check if file exists in knowledge set by name
            existing_file = None
            for file_id in file_ids:
                file = await file_repo.get_file_by_id(file_id)
                if file and file.original_filename == filename and not file.is_deleted:
                    existing_file = file
                    break

            if existing_file:
                existing_file.storage_key = new_key
                existing_file.file_size = file_size_bytes
                existing_file.content_type = content_type
                existing_file.updated_at = datetime.now(timezone.utc)
                db.add(existing_file)
                await db.commit()
                return {"success": True, "message": f"Updated file: {filename}"}
            else:
                new_file = FileCreate(
                    user_id=user_id,
                    folder_id=None,
                    original_filename=filename,
                    storage_key=new_key,
                    file_size=file_size_bytes,
                    content_type=content_type,
                    scope=FileScope.PRIVATE,
                    category=FileCategory.DOCUMENT,
                )
                created = await file_repo.create_file(new_file)
                await ks_repo.link_file_to_knowledge_set(created.id, knowledge_set_id)
                await db.commit()
                return {"success": True, "message": f"Created file: {filename}"}
    except Exception as e:
        logger.exception("persist_xlsx failed")
        return {"error": f"Internal error: {e!s}", "success": False}
